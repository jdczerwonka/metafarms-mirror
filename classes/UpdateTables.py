import os
import sys
import datetime
import time
import math
import subprocess
import logging

import simplejson as json
from requests import post as httpPost

from sqlalchemy import create_engine, func
from sqlalchemy.orm import sessionmaker

from numpy.core import defchararray
from numpy import array as nparray
import pandas
from openpyxl import load_workbook
from xlrd import open_workbook, xldate_as_tuple

from MetaFarms import MetaFarms as mf
from Tables import *

DOWNLOAD_EXT_PATH = "\\downloads"
UPLOAD_EXT_PATH = "\\uploads"
# DOWNLOAD_EXT_PATH = "\\initialize\\downloads"
# UPLOAD_EXT_PATH = "\\initialize\\uploads"
ERROR_EXT_PATH = "\\log_uploads"
LOG_EXT_PATH = "\\log_errors"
CLOSEOUT_EXT_PATH = "\\closeouts"
CLOSEOUT_STRUC_PATH = "\\closeout_structure.csv"

class UpdateTables():
    def __init__(self, Server, Database, Username, Password, CFID, urlSlack, ProgramPath, GitHubPath, bcpPath, LogName, DownloadDict, SlackPostBool = True):
        self.server = Server
        self.database = Database
        self.username = Username
        self.password = Password

        self.slack_url = urlSlack
        self.slack_post_bool = SlackPostBool

##        self.db_uri = 'mssql+pyodbc://' + self.username + ':' + self.password + '@' + self.server + '/' + self.database + '?driver=SQL+Server+Native+Client+11.0'
        self.db_uri = 'mssql+pymssql://' + self.username + ':' + self.password + '@' + self.server + '/' + self.database + '?charset=utf8'

        self.cfid = CFID

        self.github_path = GitHubPath
        self.program_path = ProgramPath
        self.bcp_path = bcpPath

        self.download_path = self.program_path + DOWNLOAD_EXT_PATH
        self.download_closeout_path = self.download_path + CLOSEOUT_EXT_PATH
        self.upload_path = self.program_path + UPLOAD_EXT_PATH
        self.error_path = self.program_path + ERROR_EXT_PATH
        self.log_path = self.program_path + LOG_EXT_PATH
        self.log_name = LogName
        self.closeout_struc_path = self.program_path + CLOSEOUT_STRUC_PATH

        logging.basicConfig(filename = self.log_path + "\\" + self.log_name + " " + datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S") + ".log", level = logging.INFO)

        self.save_path_dict = { 'ingredients' : '\\ingredients.csv',
                                'groups' : '\\groups.csv',
                                'deaths' : '\\deaths.csv',
                                'movements' : '\\movements.csv',
                                'sales' : '\\sales.csv',
                                'budget_actuals' : '\\budget_actuals.csv',
                                'closeouts' : '\\closeouts.csv'}

        self.open_path_dict = { 'ingredients' : '\\ingredients.xlsx',
                                'groups' : '\\groups.xls',
                                'deaths' : '\\deaths.xls',
                                'movements' : '\\movements.xls',
                                'sales' : '\\sales.xls',
                                'budget_actuals' : '',
                                'closeouts' : ''}

        self.error_path_dict = {'ingredients' : '\\ingredients-error.txt',
                                'groups' : '\\groups-error.txt',
                                'deaths' : '\\deaths-error.txt',
                                'movements' : '\\movements-error.txt',
                                'sales' : '\\sales-error.txt',
                                'budget_actuals' : '\\budget_actuals-error.txt',
                                'closeouts' : '\\closeouts-error.txt'}

        self.sheet_name_dict = {'ingredients' : '',
                                'groups' : 'Report',
                                'deaths' : 'Mortality Report',
                                'movements' : 'Summary',
                                'sales' : '',
                                'budget_init' : 'Budgets',
                                'budget_actuals' : '',
                                'closeouts' : 'Closeout Report'}

        # self.save_path_dict = { 'ingredients' : '\\ingredients.csv',
        #                         'groups' : '\\groups.csv',
        #                         'deaths' : '\\deaths.csv',
        #                         'movements' : '\\movements.csv',
        #                         'sales 1' : '\\sales_1.csv',
        #                         'sales 2' : '\\sales_2.csv',
        #                         'sales 3' : '\\sales_3.csv',
        #                         'sales 4' : '\\sales_4.csv',
        #                         'sales 5' : '\\sales_5.csv',
        #                         'sales 6' : '\\sales_6.csv',
        #                         'budget_init' : '\\budgets.csv',                                                                                     
        #                         'budget_actuals' : '\\budget_actuals.csv',
        #                         'closeouts' : '\\closeouts.csv'}

        # self.error_path_dict = {'ingredients' : '\\ingredients-error.txt',
        #                         'groups' : '\\groups-error.txt',
        #                         'deaths' : '\\deaths-error.txt',
        #                         'movements' : '\\movements-error.txt',
        #                         'sales 1' : '\\sales-error.txt',
        #                         'sales 2' : '\\sales-error.txt',
        #                         'sales 3' : '\\sales-error.txt',
        #                         'sales 4' : '\\sales-error.txt',
        #                         'sales 5' : '\\sales-error.txt',
        #                         'sales 6' : '\\sales-error.txt',
        #                         'budget_init' : '\\budget_init-error.txt',
        #                         'budget_actuals' : '\\budget_actuals-error.txt',
        #                         'closeouts' : '\\closeouts-error.txt'}

        # self.open_path_dict = { 'ingredients' : '\\Initialize Diets.xlsx',
        #                         'groups' : '\\groups.xlsx',
        #                         'deaths' : '\\Initialize Deaths.xlsx',
        #                         'movements' : '\\Initialize Movements.xlsx',
        #                         'sales 1' : '\\Initialize Sales 1.xls',
        #                         'sales 2' : '\\Initialize Sales 2.xls',
        #                         'sales 3' : '\\Initialize Sales 3.xls',
        #                         'sales 4' : '\\Initialize Sales 4.xls',
        #                         'sales 5' : '\\Initialize Sales 5.xls',
        #                         'sales 6' : '\\Initialize Sales 6.xls',
        #                         'budget_init' : '\\Initialize Budgets.xlsx',                                                                                                                              
        #                         'budget_actuals' : '',
        #                         'closeouts' : ''}

        self.download_dict = DownloadDict
        self.start_date = {}
        self.end_date = datetime.date.today()

        for key in self.download_dict:
            self.start_date[key] = (self.end_date - datetime.timedelta(days=self.download_dict[key]))

        self.mf = mf(self.cfid, self.download_path, self.github_path)
        
    def update(self, download_bool = True):
        try:
            if download_bool:
                self.delete_files
                self.download_files
            self.create_uploads
            self.update_tables
        except Exception:
            self.mf.close()
            logging.exception("Error!")

            if self.slack_post_bool:
                payload = {'text': 'Error logged at ' + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + ' for ' + self.log_name}
                headers = {'content-type': 'application/json'}

                r = httpPost(self.slack_url, data=json.dumps(payload), headers=headers)

            print "Error occurred"
            raise

        logging.info(self.log_name + " ran successfully at " + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        print "Run successful"

    @property
    def download_files(self):
        self.mf.open()

        for key in self.download_dict:
            if key == 'ingredients':
                self.mf.getDietIngredientDetail(self.start_date[key].strftime("%m/%d/%Y"), self.end_date.strftime("%m/%d/%Y"), 'All Feed Mills')

            elif key == 'groups':
                self.mf.getGroupList('producer', 'All Producers')

            elif key == 'movements':
                self.mf.getMovementReportSingleRow('producer', 'All Producers', self.start_date[key].strftime("%m/%d/%Y"), self.end_date.strftime("%m/%d/%Y"))

            elif key == 'deaths':
                self.mf.getMortalityList('producer', 'All Producers', self.start_date[key].strftime("%m/%d/%Y"), self.end_date.strftime("%m/%d/%Y"))

            elif key == 'sales':
                self.mf.getMarketSalesSummary('producer', 'All Producers', self.start_date[key].strftime("%m/%d/%Y"), self.end_date.strftime("%m/%d/%Y"), ['jbs', 'ipc'])

            elif key == 'budget_actuals':
                pass

            elif key == 'closeouts':
                session = self.create_session()
                query_results = session.query(Groups.group_num).filter(Groups.open_date >= self.start_date[key]).order_by(Groups.group_num.asc()).all()
                session.close()
                self.closeout_groups = nparray(query_results).flatten()
                self.closeout_groups_red = self.closeout_groups[defchararray.find(defchararray.upper(self.closeout_groups), "GEB") == -1]
                
                self.mf.getGroupDetailCloseout(self.closeout_groups_red)

        logging.info(key + ' successfully downloaded')
        self.mf.close()

    @property
    def create_uploads(self):
        for key in self.download_dict:
            if key == 'ingredients':
                self.ingredients(key)

            elif key == 'groups':
                self.groups(key)

            elif key == 'movements':
                self.movements(key)

            elif key == 'deaths':
                self.deaths(key)

            elif key == 'sales':
                self.sales(key)

            # elif key == 'sales 1' or key == 'sales 2' or key == 'sales 3' or key == 'sales 4' or key == 'sales 5' or key == 'sales 6':
            #     self.sales(key)
            elif key == 'budget_init':
                self.budget_init(key)

            elif key == 'budget_actuals':
                self.budget_actuals(key)

            elif key == 'closeouts':
                self.closeouts(key)

    @property
    def update_tables(self):
        session = self.create_session()

        for key in self.download_dict:
            if key == 'ingredients':
                session.query(Ingredients).filter(Ingredients.delivery_date >= self.start_date[key]).delete()
                session.commit()
                logging.info(key + ' successfully deleted')

                subprocess.call(self.bcp_str('ingredients', key))
                logging.info(key + ' successfully updated')

            elif key == 'groups':
                pass

            elif key == 'movements':
                session.query(Movements).filter(( Movements.movement_date >= self.start_date[key] ) & ( ~ Movements.event_category.in_(['Death', 'Euthanize']) )).delete(synchronize_session=False)
                session.commit()
                logging.info(key + ' successfully deleted')
                
                subprocess.call(self.bcp_str('movements', key))
                logging.info(key + ' successfully updated')

            elif key == 'deaths':
                session.query(Movements).filter(( Movements.movement_date >= self.start_date[key] ) & ( Movements.event_category.in_(['Death', 'Euthanize']) )).delete(synchronize_session=False)
                session.commit()
                logging.info(key + ' successfully deleted')

                subprocess.call(self.bcp_str('movements', key))
                logging.info(key + ' successfully updated')

            elif key == 'sales':
                session.query(Sales).filter(Sales.kill_date >= self.start_date[key]).delete()
                session.commit()
                logging.info(key + ' successfully deleted')
                
                subprocess.call(self.bcp_str('sales', key))
                logging.info(key + ' successfully updated')

            # elif key == 'sales 1' or key == 'sales 2' or key == 'sales 3' or key == 'sales 4' or key == 'sales 5' or key == 'sales 6':
            #     subprocess.call(self.bcp_str('sales', key))
            #     logging.info(key + ' successfully updated')
            elif key == 'budget_init':
                subprocess.call(self.bcp_str('budgets', key))
                logging.info(key + ' successfully updated')                

            elif key == 'budget_actuals':
                session.query(Budgets).filter( (Budgets.start_date >= self.start_date[key]) & (Budgets.budget_type == 'actual') ).delete()
                session.commit()
                logging.info(key + ' successfully deleted')
                
                subprocess.call(self.bcp_str('budgets', key))
                logging.info(key + ' successfully updated')

            elif key == 'closeouts':
                session.query(Closeouts).filter(Closeouts.start_date >= self.start_date[key]).delete()
                session.commit()
                logging.info(key + ' successfully deleted')
                
                subprocess.call(self.bcp_str('closeouts', 'closeouts'))
                logging.info(key + ' successfully updated')

        session.close()

    @property
    def delete_files(self):
        for file in os.listdir(self.download_closeout_path):
            os.remove(self.download_closeout_path + "\\" + file)

        for file in os.listdir(self.download_path):
            if os.path.isfile(self.download_path + "\\" + file):
                os.remove(self.download_path + "\\" + file)

        logging.info('downloads deleted')

        for file in os.listdir(self.upload_path):
            os.remove(self.upload_path + "\\" + file)

        logging.info('uploads deleted')

    def create_session(self):
        engine = create_engine(self.db_uri)
        Base.metadata.bind = engine
        DBSession = sessionmaker(bind=engine)
        session = DBSession()
        return session

    def bcp_str(self, TableStr, ReportStr):
        return self.bcp_path + " " + self.database + ".dbo." + TableStr + " in " + self.save_path(ReportStr) + " -c -U " + self.username + " -S " + self.server + " -t \\t -P " + self.password + " -e " + self.error_file_path(ReportStr)

    def save_path(self, ReportStr):
        return self.upload_path + self.save_path_dict[ReportStr]

    def open_path(self, ReportStr):
        return self.download_path + self.open_path_dict[ReportStr]

    def error_file_path(self, ReportStr):
        return self.error_path + self.error_path_dict[ReportStr]

    # def diets(self):
    #     df = pandas.read_excel(self.open_path(key), sheetname="Ingredient Quantities", skiprows=18, parse_cols=(0, 3, 6, 9, 10, 11, 14))
    #     df.columns = ['group_num', 'delivery_date', 'diet', 'quantity', 'mill', 'order_id', 'cost']

    #     df['year'] = pandas.DatetimeIndex(df['delivery_date']).year
    #     df['month'] = pandas.DatetimeIndex(df['delivery_date']).month
    #     df['week'] = pandas.DatetimeIndex(df['delivery_date']).week
    #     df['delivery_month'] = df['year'].astype(str) + " " + df['month'].astype(str)
    #     df['delivery_week'] = (df['year'] - ((df['week'] == 53) & (df['month'] == 1))).astype(str) + " " + df['week'].astype(str)

    #     df['id'] = ""
    
    #     df = df[['id', 'order_id', 'group_num', 'mill', 'delivery_date', 'delivery_month', 'delivery_week', 'year', 'month', 'week', 'diet', 'quantity', 'cost']]

    #     df.to_csv(self.save_path(key), sep='\t', index=False, header=False)
    #     logging.info(key + ' successfully parsed')

    def ingredients(self, key):
        ADD_COL = 5

        wb = load_workbook(self.open_path(key), read_only=True, use_iterators=True)
        ws = wb['Ingredient Quantities']
        max_col = ws.max_column
        wb._archive.close()

        col_arr = range(20, max_col)
        col_arr.insert(0, 11)
        col_arr.insert(0, 10)
        col_arr.insert(0, 6)
        col_arr.insert(0, 3)
        col_arr.insert(0, 0)
        


        qty_df = pandas.read_excel(self.open_path(key), sheetname="Ingredient Quantities", skiprows=18, parse_cols=col_arr)
        col_names = range(0, max_col - 20 + ADD_COL)

        col_names[0:ADD_COL] = qty_df.columns[0:ADD_COL]
        for x in range(ADD_COL, max_col - 20 + ADD_COL):
            s = qty_df.columns[x]
            col_names[x] = s[:s.find('(') - 1]

        qty_df.columns = col_names
        qty_df.rename(columns={"Order ID" : "order_id", "Group" : "group_num", "Diet" : "diet", "Delivery Date" : "delivery_date", "Mill" : "mill"}, inplace=True)
        qty_df_melt = pandas.melt(qty_df, id_vars = ['order_id', 'group_num', 'diet', 'delivery_date', 'mill'], var_name = 'ingredient', value_name='quantity')
        qty_df_melt = qty_df_melt.dropna()
        qty_df_melt['id'] = qty_df_melt['order_id'] + qty_df_melt['ingredient']
        qty_df_melt = qty_df_melt.groupby(['id', 'order_id', 'group_num', 'delivery_date', 'mill', 'diet', 'ingredient'], as_index=False).sum()

        

        cost_df = pandas.read_excel(self.open_path('ingredients'), sheetname="Ingredient Cost", skiprows=18, parse_cols=col_arr)
        col_names = range(0, max_col - 20 + ADD_COL)

        col_names[0:ADD_COL] = cost_df.columns[0:ADD_COL]
        for x in range(ADD_COL, max_col - 20 + ADD_COL):
            s = cost_df.columns[x]
            if s[len(s) - 2:] == '.1' or s[len(s) - 2:] == '.2' or s[len(s) - 2:] == '.3':
                col_names[x] = s[:len(s) - 2]
            else:
                col_names[x] = s

        cost_df.columns = col_names
        cost_df.rename(columns={"Order ID" : "order_id", "Group" : "group_num", "Diet" : "diet", "Delivery Date" : "delivery_date", "Mill" : "mill"}, inplace=True)
        cost_df_melt = pandas.melt(cost_df, id_vars = ['order_id', 'group_num', 'diet', 'delivery_date', 'mill'], var_name = 'ingredient', value_name='cost')
        cost_df_melt = cost_df_melt.dropna()
        cost_df_melt['id'] = cost_df_melt['order_id'] + cost_df_melt['ingredient']
        cost_df_melt = cost_df_melt.groupby(['id', 'order_id', 'group_num', 'delivery_date', 'mill', 'diet', 'ingredient'], as_index=False).sum()
        


        df = pandas.merge(qty_df_melt, cost_df_melt[['id', 'cost']], on='id')

        df['year'] = df['delivery_date'].dt.year
        df['month'] = df['delivery_date'].dt.month.astype(str).str.zfill(2)
        df['week'] = df['delivery_date'].dt.week.astype(str).str.zfill(2)
        df['delivery_month'] = df['year'].astype(str) + ' ' + df['month']
        df['delivery_week'] = (df['year'] - ((df['week'] == '53') & (df['month'] == '01'))).astype(str) + ' ' + df['week']

        df['quantity'] = df['quantity'].astype(float).round(2)
        df['cost'] = df['cost'].astype(float).round(2)
        
        df['id'] = ""
        df = df[['id', 'order_id', 'group_num', 'mill', 'delivery_date', 'delivery_month', 'delivery_week', 'year', 'month', 'week', 'diet', 'ingredient', 'quantity', 'cost']]

        df.to_csv(self.save_path(key), sep='\t', index=False, header=False)
        logging.info(key + ' successfully parsed')

    def groups(self, key):
        excel_df = pandas.read_excel(self.open_path(key), sheetname=self.sheet_name_dict[key], skiprows=9, parse_cols=(14, 15, 16, 19, 23, 24, 25, 26))
        excel_df.columns = ['producer', 'site', 'barn', 'group_num', 'group_type', 'status', 'open_date', 'close_date']
        excel_df = excel_df[['group_num', 'group_type', 'status', 'producer', 'site', 'barn', 'open_date', 'close_date']]

        engine = create_engine(self.db_uri)
        sql_df = pandas.read_sql('groups', engine)

        df_comb = excel_df.append(sql_df)
        group_num_unique = df_comb[~df_comb.duplicated(subset='group_num', keep=False)]['group_num']

        group_num_remove = sql_df[sql_df['group_num'].isin(group_num_unique)]['group_num']
        add_df = excel_df[excel_df['group_num'].isin(group_num_unique)]

        group_num_change_df = df_comb[~df_comb.duplicated(subset=['group_num', 'status', 'open_date', 'close_date'], keep=False)]
        group_num_change = group_num_change_df[group_num_change_df.duplicated(subset='group_num')]['group_num']

        change_df = excel_df[excel_df['group_num'].isin(group_num_change)][['group_num', 'status', 'open_date', 'close_date']]
        change_df = change_df.fillna("")
        logging.info(key + ' successfully parsed')


        
        session = self.create_session()
        
        if not group_num_unique.empty:
            session.query(Groups).filter(Groups.group_num.in_(group_num_remove)).delete(synchronize_session=False)
            session.commit()
            logging.info(key + ' successfully deleted')
                
        if not change_df.empty:
            for index, row in change_df.iterrows():
                session.query(Groups).filter(Groups.group_num == row['group_num']).update({'status' : row['status'], 'open_date' : row['open_date'], 'close_date' : row['close_date']})
                session.commit()
            logging.info(key + ' successfully updated')

        if not add_df.empty:
            add_df.to_sql('groups', engine, flavor='mssql', if_exists='append', index=False)
            logging.info(key + ' successfully added')

        session.close()

    def budget_init(self, key):
        df = pandas.read_excel(self.open_path(key), sheetname=self.sheet_name_dict[key])
        df.to_csv(self.save_path(key), sep='\t', index=False, header=False)
        logging.info(key + ' successfully parsed')

    def budget_actuals(self, key):
        session = self.create_session()
        query_results = session.query(Groups.group_num).filter(Groups.open_date >= self.start_date[key]).filter(Groups.producer != "Mark Gebben").filter((Groups.group_type == "Finishing") | (Groups.group_type == "Wean to Finish")).all()
        df = pandas.DataFrame(query_results)
        df['budget_type'] = 'actual'
        df['budget_id'] = df['group_num'] + '-' + df['budget_type']
        df = df[['budget_id', 'group_num', 'budget_type']]

        query_results = session.query(Closeouts.group_num, Closeouts.group_type, Closeouts.producer, Closeouts.site, Closeouts.barn, Closeouts.start_date, Closeouts.close_date,
            Closeouts.avg_dof, Closeouts.total_location_days, Closeouts.number_pigs_started, Closeouts.avg_wt_pigs_started,
            Closeouts.condemned_market_hog_out_number, Closeouts.dead_on_transport_out_number, Closeouts.deaths_out_number, Closeouts.discount_market_hog_out_number,
            Closeouts.gilt_arrival_to_iso_out_number, Closeouts.gilt_sale_external_out_number, Closeouts.no_value_out_number, Closeouts.standard_market_sale_out_number,
            Closeouts.sow_arrival_out_number, Closeouts.transfer_feeder_pig_out_number, Closeouts.yard_dead_out_number,
            Closeouts.condemned_market_hog_out_total_wt, Closeouts.dead_on_transport_out_total_wt, Closeouts.deaths_out_total_wt, Closeouts.discount_market_hog_out_total_wt,
            Closeouts.gilt_arrival_to_iso_out_total_wt, Closeouts.gilt_sale_external_out_total_wt, Closeouts.no_value_out_total_wt, Closeouts.standard_market_sale_out_total_wt,
            Closeouts.sow_arrival_out_total_wt, Closeouts.transfer_feeder_pig_out_total_wt, Closeouts.yard_dead_out_total_wt,
            Closeouts.condemned_market_hog_out_receipts, Closeouts.dead_on_transport_out_receipts, Closeouts.deaths_out_receipts, Closeouts.discount_market_hog_out_receipts,
            Closeouts.gilt_arrival_to_iso_out_receipts, Closeouts.gilt_sale_external_out_receipts, Closeouts.no_value_out_receipts, Closeouts.standard_market_sale_out_receipts,
            Closeouts.sow_arrival_out_receipts, Closeouts.transfer_feeder_pig_out_receipts, Closeouts.yard_dead_out_receipts,
            Closeouts.fcr_w_dead, Closeouts.totals_summary_pigs_in_pig
            ).filter(Closeouts.group_num.in_(df['group_num'])).all()
        
        closeout_df = pandas.DataFrame(query_results)

        closeout_df['market_num'] = closeout_df['standard_market_sale_out_number']
        closeout_df['market_avg_wt'] = ( ( closeout_df['standard_market_sale_out_total_wt'] ).astype(float) / closeout_df['market_num'].astype(float) ).round(1)
        closeout_df['market_avg_price_cwt'] = ( ( closeout_df['standard_market_sale_out_receipts'] ).astype(float) / closeout_df['market_num'].astype(float) ).round(2)

        closeout_df['discount_num'] = closeout_df['discount_market_hog_out_number'] + closeout_df['no_value_out_number'] + closeout_df['condemned_market_hog_out_number']
        closeout_df['discount_avg_wt'] = ( ( closeout_df['discount_market_hog_out_total_wt'] + closeout_df['no_value_out_total_wt'] + closeout_df['condemned_market_hog_out_total_wt'] ).astype(float) / closeout_df['discount_num'].astype(float) ).round(1)
        closeout_df['discount_avg_price_pig'] = ( ( closeout_df['discount_market_hog_out_total_wt'] + closeout_df['no_value_out_total_wt'] + closeout_df['condemned_market_hog_out_total_wt'] ).astype(float) / closeout_df['discount_num'].astype(float) ).round(2)

        closeout_df['death_num'] = closeout_df['deaths_out_number']
        closeout_df['death_avg_wt'] = ( ( closeout_df['deaths_out_total_wt'] ) / closeout_df['death_num'].astype(float) ).round(1)

        closeout_df['transport_death_num'] = closeout_df['yard_dead_out_number'] + closeout_df['dead_on_transport_out_number']
        closeout_df['transport_death_avg_wt'] = ( ( closeout_df['yard_dead_out_total_wt'] + closeout_df['dead_on_transport_out_total_wt'] ).astype(float) / closeout_df['transport_death_num'].astype(float) ).round(1)

        closeout_df['gilt_num'] = closeout_df['gilt_arrival_to_iso_out_number'] + closeout_df['gilt_sale_external_out_number'] + closeout_df['sow_arrival_out_number']
        closeout_df['gilt_avg_wt'] = ( ( closeout_df['gilt_arrival_to_iso_out_total_wt'] + closeout_df['gilt_sale_external_out_total_wt'] + closeout_df['sow_arrival_out_total_wt'] ).astype(float) / closeout_df['gilt_num'].astype(float) ).round(1)
        closeout_df['gilt_avg_price_pig'] = ( ( closeout_df['gilt_arrival_to_iso_out_receipts'] + closeout_df['gilt_sale_external_out_receipts'] + closeout_df['sow_arrival_out_receipts'] ).astype(float) / closeout_df['gilt_num'].astype(float) ).round(2)

        closeout_df['transfer_num'] = closeout_df['transfer_feeder_pig_out_number']
        closeout_df['transfer_avg_wt'] = ( ( closeout_df['transfer_feeder_pig_out_total_wt'] ).astype(float) / closeout_df['transfer_num'].astype(float) ).round(1)
        closeout_df['transfer_avg_price_pig'] = ( ( closeout_df['transfer_feeder_pig_out_receipts'] ).astype(float) / closeout_df['transfer_num'].astype(float) ).round(2)
        
        closeout_df['wof_avg'] = (closeout_df['avg_dof'].astype(float) / 7).round(1)
        closeout_df['wof_tot'] = (closeout_df['total_location_days'].astype(float) / 7).round(1)

        closeout_df.rename(columns = {'number_pigs_started' : 'start_num', 'avg_wt_pigs_started' : 'start_avg_wt', 'fcr_w_dead' : 'feed_conversion', 'totals_summary_pigs_in_pig' : 'pig_cost_pig'}, inplace = True)

        df = df.merge(closeout_df[[ 'group_num', 'group_type', 'producer', 'site', 'barn', 
                                    'start_date', 'close_date', 'wof_avg', 'wof_tot', 'start_num', 'start_avg_wt',
                                    'market_num', 'gilt_num', 'discount_num', 'death_num', 'transport_death_num', 'transfer_num',
                                    'market_avg_wt', 'gilt_avg_wt', 'discount_avg_wt', 'death_avg_wt', 'transport_death_avg_wt', 'transfer_avg_wt',
                                    'market_avg_price_cwt', 'gilt_avg_price_pig', 'discount_avg_price_pig', 'transfer_avg_price_pig',
                                    'feed_conversion']],
                                    on = 'group_num', how = 'left')
        
        query_results = session.query(Sales.group_num, func.avg(Sales.yield_per).label('yield_per'), func.avg(Sales.lean_per).label('lean_per'), func.stdev(Sales.avg_carcass_wt).label('market_std_dev_carcass_wt')).filter(Sales.group_num.in_(df['group_num'])).group_by(Sales.group_num).all()
        market_df = pandas.DataFrame(query_results)
        market_df.iloc[:, 1:] = market_df.iloc[:, 1:].astype(float).round(4)

        df = df.merge(market_df, on = 'group_num', how = 'left')

        a_query = session.query(Sales.load_id, Sales.group_num, func.stdev(Sales.avg_carcass_wt).label('load_std_dev_carcass_wt')).filter(Sales.group_num.in_(df['group_num'])).group_by(Sales.load_id, Sales.group_num).order_by(Sales.group_num).subquery()
        query_results = session.query(a_query.columns.group_num, func.avg(a_query.columns.load_std_dev_carcass_wt).label('load_avg_std_dev_carcass_wt')).group_by(a_query.columns.group_num).all()
        market_df = pandas.DataFrame(query_results)
        market_df.iloc[:, 1:] = market_df.iloc[:, 1:].astype(float).round(4)

        df = df.merge(market_df, on = 'group_num', how = 'left')

        query_results = session.query(Budgets.group_num, Budgets.feed_grinding_rate, Budgets.feed_delivery_rate, Budgets.trucking_market_rate, Budgets.trucking_feeder_rate, Budgets.rent_cost_week, Budgets.overhead_cost_pig).filter( (Budgets.group_num.in_(df['group_num'])) & (Budgets.budget_type == 'budget') ).all()
        budget_df = pandas.DataFrame(query_results)

        df = df.merge(budget_df[['group_num', 'feed_grinding_rate', 'feed_delivery_rate']], on = 'group_num', how = 'left')

        a_query = session.query(Ingredients.order_id, Ingredients.group_num, func.sum(Ingredients.weight).label('weight')).filter( (Ingredients.group_num.in_(df['group_num'])) & (Ingredients.mill == 'Home Mill') ).group_by(Ingredients.order_id, Ingredients.group_num).subquery()
        query_results = session.query(a_query.columns.group_num, func.sum(a_query.columns.weight).label('total_wt'), func.count(a_query.columns.order_id).label('trips')).group_by(a_query.columns.group_num).all()
        diet_df = pandas.DataFrame(query_results)

        diet_df['avg_feed_delivery_ton'] = ( diet_df['total_wt'].astype(float) / diet_df['trips'].astype(float) / 2000).round(1)
        df = df.merge(diet_df[['group_num', 'avg_feed_delivery_ton']], on = 'group_num', how = 'left')

        df = df.merge(budget_df[['group_num', 'trucking_market_rate']], on = 'group_num', how = 'left')

        query_results = session.query(Movements.group_num, func.avg(Movements.quantity).label('avg_trucking_market_pig')).filter( (Movements.group_num.in_(df['group_num'])) & (Movements.event_category == 'Market Sale') & (Movements.movement_type == 'out') ).group_by(Movements.group_num).all()
        movement_df = pandas.DataFrame(query_results)

        df = df.merge(movement_df, on = 'group_num', how = 'left')

        df = df.merge(budget_df[['group_num', 'trucking_feeder_rate']], on = 'group_num', how = 'left')

        query_results = session.query(Movements.group_num, func.avg(Movements.quantity).label('avg_trucking_feeder_pig')).filter( (Movements.group_num.in_(df['group_num'])) & (Movements.event_category != 'Inventory Adjustment') & (Movements.movement_type == 'in') ).group_by(Movements.group_num).all()
        movement_df = pandas.DataFrame(query_results)

        df = df.merge(movement_df, on = 'group_num', how = 'left')

        df = df.merge(budget_df[['group_num', 'rent_cost_week']], on = 'group_num', how = 'left')

        df = df.merge(closeout_df[['group_num', 'pig_cost_pig']], on = 'group_num', how = 'left')

        df = df.merge(budget_df[['group_num', 'overhead_cost_pig']], on = 'group_num', how = 'left')

        df.fillna(0, inplace = True)
        df.loc[df['close_date'] == 0, 'close_date'] = ''

        df.to_csv(self.save_path(key), sep='\t', index=False, header=False)
        logging.info(key + ' successfully parsed')             
        session.close()

    def deaths(self, key):
        df = pandas.read_excel(self.open_path(key), sheetname=self.sheet_name_dict[key], skiprows=7, parse_cols=(8, 11, 13, 14, 15, 16))
        df.columns = ['group_num', 'movement_date', 'quantity', 'weight', 'event_category', 'event_name']

        df['year'] = df['movement_date'].dt.year
        df['month'] = df['movement_date'].dt.month.astype(str).str.zfill(2)
        df['week'] = df['movement_date'].dt.week.astype(str).str.zfill(2)
        df['movement_month'] = df['year'].astype(str) + ' ' + df['month']
        df['movement_week'] = (df['year'] - ((df['week'] == '53') & (df['month'] == '01'))).astype(str) + ' ' + df['week']

        df['movement_type'] = 'out'
        df['entity_type'] = 'group'
        df['event_code'] = ''
        df['cost'] = 0
        df['movement_id'] = ''
        df['id'] = ''

        df = df[['id', 'movement_id', 'movement_type', 'group_num', 'entity_type', 'event_category', 'event_code', 'event_name', 'movement_date', 'movement_month', 'movement_week', 'year', 'month', 'week', 'quantity', 'weight', 'cost']]

        df.to_csv(self.save_path(key), sep='\t', index=False, header=False)
        logging.info(key + ' successfully parsed')

    def movements(self, key):
        df = pandas.read_excel(self.open_path(key), sheetname=self.sheet_name_dict[key], skiprows=9, parse_cols=(2, 6, 10, 12, 15, 20, 21, 22, 23, 26, 28, 29, 31))
        df.columns = ['sow_unit', 'group', 'supplier', 'customer', 'plant', 'event_category', 'event_code', 'event_name', 'movement_id', 'movement_date', 'quantity', 'weight', 'cost']

        df.loc[~ df['group'].isnull(), 'sow_unit'] = float('NaN')

        df = pandas.melt(df, id_vars=['event_category', 'event_code', 'event_name', 'movement_id', 'movement_date', 'quantity', 'weight', 'cost'], value_vars=['sow_unit', 'group', 'supplier', 'customer', 'plant'], var_name='entity_type', value_name='group_num')
        df = df[~ df['group_num'].isnull()]

        df['movement_type'] = 'in'
        df.loc[df['quantity'] < 0, 'movement_type'] = 'out'
        df.loc[df['group_num'] == 'Metafarms setup existing wean to finish groups', 'group_num'] = 'Metafarms setup'

        df['quantity'] = df['quantity'].abs()
        df['weight'] = df['weight'].abs()
        df['cost'] = df['cost'].abs()

        df['year'] = df['movement_date'].dt.year
        df['month'] = df['movement_date'].dt.month.astype(str).str.zfill(2)
        df['week'] = df['movement_date'].dt.week.astype(str).str.zfill(2)
        df['movement_month'] = df['year'].astype(str) + ' ' + df['month']
        df['movement_week'] = (df['year'] - ((df['week'] == '53') & (df['month'] == '01'))).astype(str) + ' ' + df['week']

        df['id'] = ""

        df = df[['id', 'movement_id', 'movement_type', 'group_num', 'entity_type', 'event_category', 'event_code', 'event_name', 'movement_date', 'movement_month', 'movement_week', 'year', 'month', 'week', 'quantity', 'weight', 'cost']]

        df.to_csv(self.save_path(key), sep='\t', index=False, header=False)
        logging.info(key + ' successfully parsed')

    def sales(self, key):
        # pandas.options.mode.chained_assignment = None  # default='warn'

        carcass_df = pandas.read_excel(self.open_path(key), sheetname="Carcass Details", skiprows=9, parse_cols=(1, 4, 6, 7, 8, 9, 10, 14, 15, 16, 20, 21, 22, 23, 24, 31))
        carcass_df.columns = ['plant', 'tattoo', 'kill_date', 'quantity', 'avg_live_wt', 'avg_carcass_wt', 'base_price_cwt', 'value_cwt', 'vob_cwt', 'base_price', 'value', 'vob', 'back_fat', 'loin_depth', 'lean', 'yield']
        carcass_df['tattoo'] = carcass_df['tattoo'].astype(int).astype(str).str.zfill(4)

        load_df = pandas.read_excel(self.open_path(key), sheetname="Sales", skiprows=21, parse_cols=(20, 28, 34, 37))
        load_df.columns = ['group_num', 'kill_date', 'plant', 'tattoo']
        load_df = load_df[load_df['tattoo'].notnull()]
        load_df['tattoo'] = load_df['tattoo'].astype(int).astype(str).str.zfill(4)

        carcass_df['load_id'] = carcass_df['tattoo'] + carcass_df['kill_date'].astype(str)
        load_df['load_id'] = load_df['tattoo'] + load_df['kill_date'].astype(str)
        load_id_dupl = load_df[load_df.duplicated(subset='load_id', keep=False)]['load_id']

        df = pandas.merge(carcass_df, load_df[['group_num', 'load_id']], on='load_id')
        df_final = df[~ df['load_id'].isin(load_id_dupl)]

        df_final['avg_carcass_wt'] = df_final['avg_carcass_wt'].round(0)
        df_final['avg_live_wt'] = df_final['avg_live_wt'].round(1)
        df_final['lean'] = df_final['lean'] * 100
        df_final['base_price_cwt'] = df_final['base_price_cwt'].round(2)
        df_final['vob_cwt'] = df_final['vob_cwt'].round(2)
        df_final['value_cwt'] = df_final['value_cwt'].round(2)
        df_final['base_price'] = df_final['base_price'].round(2)
        df_final['vob'] = df_final['vob'].round(2)
        df_final['value'] = df_final['value'].round(2)

        df_final['year'] = df_final['kill_date'].dt.year
        df_final['month'] = df_final['kill_date'].dt.month.astype(str).str.zfill(2)
        df_final['week'] = df_final['kill_date'].dt.week.astype(str).str.zfill(2)
        df_final['kill_month'] = df_final['year'].astype(str) + ' ' + df_final['month']
        df_final['kill_week'] = (df_final['year'] - ((df_final['week'] == '53') & (df_final['month'] == '01'))).astype(str) + ' ' + df_final['week']

        df_final['load_id'] = df_final['group_num'] + '-' + df_final['tattoo']
        df_final['id'] = ""
        df_final = df_final[['id', 'group_num', 'load_id', 'plant', 'tattoo', 'kill_date', 'kill_month', 'kill_week', 'year', 'month', 'week', 'quantity', 'avg_live_wt', 'avg_carcass_wt', 'base_price_cwt', 'vob_cwt', 'value_cwt', 'base_price', 'vob', 'value', 'back_fat', 'loin_depth', 'yield', 'lean']]

        df_final.to_csv(self.save_path(key), sep='\t', index=False, header=False)
        logging.info(key + ' successfully parsed')

    def closeouts(self, key):
        CATEGORY_ROW_ARR = ["Group Detail Report", "Summary Pigs IN", "Summary Pigs OUT", "Adjustments", "Performance",
                            "Feed Cost Summary", "P & L Summary", "Expenses", "Group Comments"]

        DF_COL_SEARCH = "Search Str"
        DF_COL_ADJ_STR = "Adjust Str"
        DF_COL_COL_STR = "Column Str"
        DF_COL_HEADER = "Header Str"
        DF_COL_FIXED = "Fixed"
        DF_COL_DATA_TYPE = "Data Type"
        DF_COL_ROUND = "Round"
        DF_COL_CATEGORY = "Category"
        DF_COL_ROW = "Row"
        DF_COL_COL = "Column"
        DF_COL_COL_VAL = "Column Value"

        struc_df = pandas.read_csv(self.closeout_struc_path, engine="c",
                                   dtype={DF_COL_FIXED : bool, DF_COL_CATEGORY : int, DF_COL_ROW : int, DF_COL_COL : int, DF_COL_COL_VAL : int})

        df = pandas.DataFrame(index = os.listdir(self.download_closeout_path), columns = struc_df[DF_COL_HEADER].values)

        df.ix[:,len(struc_df[struc_df[DF_COL_CATEGORY] == 0]):] = 0

        for group in os.listdir(self.download_closeout_path):
        ##group = "1505BAHL.xls"

            wb = open_workbook(self.download_closeout_path + "\\" + group, on_demand=True)
            ws = wb.sheet_by_name(self.sheet_name_dict[key])

            for index, row in struc_df[struc_df[DF_COL_FIXED]].iterrows():
                if ws.cell(row[DF_COL_ROW], row[DF_COL_COL]).value != "":
                    if row[DF_COL_DATA_TYPE] == "date":
                        value = xldate_as_tuple(ws.cell(row[DF_COL_ROW], row[DF_COL_COL]).value, 0)
                        value = datetime.date(value[0], value[1], value[2])
                    elif row[DF_COL_DATA_TYPE] == "split":
                        value = str(ws.cell(row[DF_COL_ROW], row[DF_COL_COL]).value)
                        value = value[value.find(":") + 2:]
                ##    elif row[DF_COL_DATA_TYPE] == "str":
                ##        value = str(ws.cell(row[DF_COL_ROW], row[DF_COL_COL]).value)
                    else:
                        value = ws.cell(row[DF_COL_ROW], row[DF_COL_COL]).value
                else:
                    value = ws.cell(row[DF_COL_ROW], row[DF_COL_COL]).value
                    
                df.set_value(group, row[DF_COL_HEADER], value)

            category_row_dict = {}
            col = ws.col(0)
            x = 0
            for category in CATEGORY_ROW_ARR:
                while col[x].value != category:
                    x = x + 1

                category_row_dict[category] = x

            for x in range(1, len(CATEGORY_ROW_ARR) - 1):
                category_df = struc_df[struc_df[DF_COL_CATEGORY] == x]
                col_num_arr = category_df[~category_df.duplicated(subset = DF_COL_COL)][DF_COL_COL]
                
                for col_num in col_num_arr:
                    col = ws.col(col_num)
                    
                    for row_num in range(category_row_dict[CATEGORY_ROW_ARR[x]] + 2, category_row_dict[CATEGORY_ROW_ARR[x + 1]] - 2):
                        if col[row_num].value != "":
                            for index, row in category_df[(category_df[DF_COL_SEARCH] == col[row_num].value) & (category_df[DF_COL_COL] == col_num)].iterrows():
                                if ws.cell(row_num, row[DF_COL_COL_VAL]).value == "" or ws.cell(row_num, row[DF_COL_COL_VAL]).value == "-" or ws.cell(row_num, row[DF_COL_COL_VAL]).value < 0:
                                    value = 0
                                else:
                                    value = ws.cell(row_num, row[DF_COL_COL_VAL]).value
                                    
                                df.set_value(group, row[DF_COL_HEADER], value)

                                
        df["totals_summary_pigs_in_number"] = 0
        for index, row in struc_df[(struc_df[DF_COL_CATEGORY] == 1) & (struc_df[DF_COL_COL_STR] == "in_number")].iterrows():
            if row[DF_COL_ADJ_STR] != "totals_summary_pigs": 
                df["totals_summary_pigs_in_number"] = df["totals_summary_pigs_in_number"] + df[row[DF_COL_HEADER]]

        df["totals_summary_pigs_in_total_wt"] = 0
        for index, row in struc_df[(struc_df[DF_COL_CATEGORY] == 1) & (struc_df[DF_COL_COL_STR] == "in_total_wt")].iterrows():
            if row[DF_COL_ADJ_STR] != "totals_summary_pigs": 
                df["totals_summary_pigs_in_total_wt"] = df["totals_summary_pigs_in_total_wt"] + df[row[DF_COL_HEADER]]

        df["totals_summary_pigs_in_receipts"] = 0
        for index, row in struc_df[(struc_df[DF_COL_CATEGORY] == 1) & (struc_df[DF_COL_COL_STR] == "in_receipts")].iterrows():
            if row[DF_COL_ADJ_STR] != "totals_summary_pigs": 
                df["totals_summary_pigs_in_receipts"] = df["totals_summary_pigs_in_receipts"] + df[row[DF_COL_HEADER]]

        for index, row in struc_df[(struc_df[DF_COL_CATEGORY] == 1) & (struc_df[DF_COL_COL_STR] == "in_number")].iterrows():
            df[row[DF_COL_ADJ_STR] + "_in_per"] = df[row[DF_COL_ADJ_STR] + "_in_number"].astype("float64") / df["totals_summary_pigs_in_number"].astype("float64")
            df[row[DF_COL_ADJ_STR] + "_in_pig"] = df[row[DF_COL_ADJ_STR] + "_in_receipts"].astype("float64") / df[row[DF_COL_HEADER]].astype("float64")
            df[row[DF_COL_ADJ_STR] + "_in_lb"] = df[row[DF_COL_ADJ_STR] + "_in_receipts"].astype("float64") / df[row[DF_COL_ADJ_STR] + "_in_total_wt"].astype("float64")
            df[row[DF_COL_ADJ_STR] + "_in_avg_wt"] = df[row[DF_COL_ADJ_STR] + "_in_total_wt"].astype("float64") / df[row[DF_COL_HEADER]].astype("float64")



        df["totals_summary_pigs_out_number"] = 0
        for index, row in struc_df[(struc_df[DF_COL_CATEGORY] == 2) & (struc_df[DF_COL_COL_STR] == "out_number")].iterrows():
            if row[DF_COL_ADJ_STR] != "totals_summary_pigs": 
                df["totals_summary_pigs_out_number"] = df["totals_summary_pigs_out_number"] + df[row[DF_COL_HEADER]]

        df["totals_summary_pigs_out_total_wt"] = 0
        for index, row in struc_df[(struc_df[DF_COL_CATEGORY] == 2) & (struc_df[DF_COL_COL_STR] == "out_total_wt")].iterrows():
            if row[DF_COL_ADJ_STR] != "totals_summary_pigs": 
                df["totals_summary_pigs_out_total_wt"] = df["totals_summary_pigs_out_total_wt"] + df[row[DF_COL_HEADER]]

        df["totals_summary_pigs_out_receipts"] = 0
        for index, row in struc_df[(struc_df[DF_COL_CATEGORY] == 2) & (struc_df[DF_COL_COL_STR] == "out_receipts")].iterrows():
            if row[DF_COL_ADJ_STR] != "totals_summary_pigs": 
                df["totals_summary_pigs_out_receipts"] = df["totals_summary_pigs_out_receipts"] + df[row[DF_COL_HEADER]]

        for index, row in struc_df[(struc_df[DF_COL_CATEGORY] == 2) & (struc_df[DF_COL_COL_STR] == "out_number")].iterrows():
            df[row[DF_COL_ADJ_STR] + "_out_per"] = df[row[DF_COL_ADJ_STR] + "_out_number"].astype("float64") / df["totals_summary_pigs_out_number"].astype("float64")
            df[row[DF_COL_ADJ_STR] + "_out_pig"] = df[row[DF_COL_ADJ_STR] + "_out_receipts"].astype("float64") / df[row[DF_COL_HEADER]].astype("float64")
            df[row[DF_COL_ADJ_STR] + "_out_lb"] = df[row[DF_COL_ADJ_STR] + "_out_receipts"].astype("float64") / df[row[DF_COL_ADJ_STR] + "_out_total_wt"].astype("float64")
            df[row[DF_COL_ADJ_STR] + "_out_avg_wt"] = df[row[DF_COL_ADJ_STR] + "_out_total_wt"].astype("float64") / df[row[DF_COL_HEADER]].astype("float64")

                
            df.fillna(0, inplace=True)
            
            for index, row in struc_df.iterrows():
                if row[DF_COL_ROUND] > 0:
                        df[row[DF_COL_HEADER]] = df[row[DF_COL_HEADER]].round(row[DF_COL_ROUND])
       

        df.to_csv(self.save_path(key), sep='\t', index=False, header=False)
        logging.info(key + ' successfully parsed')

    def drop_table(self, table):
        engine = create_engine(self.db_uri)
        Base.metadata.tables[table.__tablename__].drop(bind=engine)

    def create_table(self, table):
        engine = create_engine(self.db_uri)
        Base.metadata.tables[table.__tablename__].create(bind=engine)
